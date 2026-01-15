from django.contrib import admin
from django.urls import path
from django.shortcuts import render
from django import forms
from django.utils import timezone
from django.db.models import Sum, Count, F, Q
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.http import HttpResponse
import csv
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from django.conf import settings
import os

User = get_user_model()

try:
    import openpyxl
    from openpyxl.styles import Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from objednavky.models import Order, OrderItem
from jidelnicek.models import DruhJidla
from .models import ReportDummy
from datetime import timedelta
from collections import defaultdict




PERIOD_CHOICES = [
    ('today', 'Dnes'),
    ('yesterday', 'Včera'),
    ('week', 'Minulý týden'),
    ('month', 'Minulý měsíc'),
    ('current_month', 'Aktuální měsíc'),
    ('year', 'Aktuální rok'),
    ('custom', 'Vlastní období'),
]

GROUPING_CHOICES = [
    ('day', 'Po dnech'),
    ('total', 'Celkem'),
]



class ReportForm(forms.Form):
    period = forms.ChoiceField(
        choices=PERIOD_CHOICES,
        initial='today',
        widget=forms.Select(attrs={'class': 'form-control'})
    )
    grouping = forms.ChoiceField(
        choices=GROUPING_CHOICES,
        initial='day',
        widget=forms.Select(attrs={'class': 'form-control'})
    )
    group = forms.ModelChoiceField(
        queryset=Group.objects.all(),
        required=False,
        empty_label='Všechny skupiny',
        widget=forms.Select(attrs={'class': 'form-control'})
    )
    customer = forms.ModelChoiceField(
        queryset=get_user_model().objects.all(),
        required=False,
        empty_label='Všichni zákazníci (vyhledávání)',
        widget=forms.Select(attrs={'class': 'form-control select2-search'})
    )
    date_from = forms.DateField(
        required=False,
        widget=forms.DateInput(attrs={'type': 'date', 'class': 'form-control'})
    )
    date_to = forms.DateField(
        required=False,
        widget=forms.DateInput(attrs={'type': 'date', 'class': 'form-control'})
    )
    # NOVÉ: Multi-select pro druhy jídel
    food_types = forms.ModelMultipleChoiceField(
        queryset=DruhJidla.objects.all(),
        required=False,
        widget=forms.SelectMultiple(attrs={'class': 'form-control', 'size': '5'}),
        label='Druhy jídel'
    )
    # NOVÉ: Fulltext search
    search = forms.CharField(
        required=False,
        widget=forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Hledat...'}),
        label='Vyhledávání'
    )



@admin.register(ReportDummy)
class ReportAdmin(admin.ModelAdmin):
    change_list_template = 'admin/reporty/dashboard.html'

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [path('', self.admin_site.admin_view(self.dashboard_view), name='report_dashboard')]
        return custom_urls + urls


    def get_report_calculations(self, form):
        """Pomocná metoda pro výpočet dat reportu sdílená mezi view a exportem"""
        today = timezone.now().date()
        report_data = []
        totals = {'unclaimed_total': 0, 'dotace': 0, 'final_price': 0}
        grouping = 'day'

        if not form.is_valid():
            return report_data, totals, grouping

        period = form.cleaned_data['period']
        grouping = form.cleaned_data.get('grouping', 'day')

        # výpočet období
        if period == 'today':
            date_from = date_to = today
        elif period == 'yesterday':
            date_from = date_to = today - timedelta(days=1)
        elif period == 'week':
            date_from = date_to = today - timedelta(days=7)
        elif period == 'month':
            date_from = date_to = today.replace(day=1) - timedelta(days=1)
        elif period == 'current_month':
            date_from = today.replace(day=1)
            date_to = today
        elif period == 'year':
            date_from = today.replace(month=1, day=1)
            date_to = today
        else:  # custom
            date_from = form.cleaned_data.get('date_from')
            date_to = form.cleaned_data.get('date_to')

        # základní queryset
        queryset = Order.objects.filter(
            status__in=['vydano', 'nevyzvednuto']
        ).select_related('user').prefetch_related('items').order_by('datum_vydeje', 'user__first_name')

        if date_from and date_to:
            queryset = queryset.filter(datum_vydeje__range=[date_from, date_to])

        if form.cleaned_data['group']:
            queryset = queryset.filter(user__groups=form.cleaned_data['group'])

        if form.cleaned_data['customer']:
            queryset = queryset.filter(user=form.cleaned_data['customer'])

        for order in queryset:
            issued_items = order.items.filter(vydano=True)
            if issued_items.exists():
                unclaimed_total = sum(
                    item.quantity * getattr(item.menu_item.jidlo, 'cena', 0) for item in issued_items
                )
                dotace = sum(
                    (getattr(item.menu_item.jidlo, 'cena', 0) - item.cena) * item.quantity
                    for item in issued_items
                )
                final_price = sum(item.quantity * item.cena for item in issued_items)

                row = {
                    'user': order.user.get_full_name(),
                    'osobni_cislo': order.user.osobni_cislo or '',
                    'identifikacni_medium': order.user.identifikacni_medium or '',
                    'date': order.datum_vydeje,
                    'status': order.status,
                    'unclaimed_total': round(unclaimed_total, 2),
                    'dotace': round(dotace, 2),
                    'final_price': round(final_price, 2),
                }
                report_data.append(row)
                totals['unclaimed_total'] += unclaimed_total
                totals['dotace'] += dotace
                totals['final_price'] += final_price

        if grouping == 'total':
            grouped = defaultdict(
                lambda: {'unclaimed_total': 0, 'dotace': 0, 'final_price': 0, 'count': 0, 'osobni_cislo': '', 'identifikacni_medium': ''}
            )
            
            for row in report_data:
                key = row['user']
                grouped[key]['unclaimed_total'] += row['unclaimed_total']
                grouped[key]['dotace'] += row['dotace']
                grouped[key]['final_price'] += row['final_price']
                grouped[key]['count'] += 1
                if not grouped[key]['osobni_cislo']:
                    grouped[key]['osobni_cislo'] = row['osobni_cislo']
                if not grouped[key]['identifikacni_medium']:
                    grouped[key]['identifikacni_medium'] = row['identifikacni_medium']
            
            report_data = [{
                'user': user,
                'osobni_cislo': data['osobni_cislo'],
                'identifikacni_medium': data['identifikacni_medium'],
                'unclaimed_total': round(data['unclaimed_total'], 2),
                'dotace': round(data['dotace'], 2),
                'final_price': round(data['final_price'], 2),
                'count': data['count'],
                'grouped': True
            } for user, data in sorted(grouped.items(), key=lambda x: x[1]['final_price'], reverse=True)]

        # NOVÉ: Fulltext search
        search_query = form.cleaned_data.get('search', '').strip()
        if search_query:
            filtered_data = []
            for row in report_data:
                searchable_text = ' '.join([
                    str(row.get('user', '')),
                    str(row.get('osobni_cislo', '')),
                    str(row.get('identifikacni_medium', '')),
                    str(row.get('date', '')),
                    str(row.get('status', '')),
                    str(row.get('unclaimed_total', '')),
                    str(row.get('dotace', '')),
                    str(row.get('final_price', '')),
                ]).lower()
                
                if search_query.lower() in searchable_text:
                    filtered_data.append(row)
            
            report_data = filtered_data

        return report_data, totals, grouping


    def get_order_items_report(self, form):
        """Report pro položky objednávek - počet a názvy jídel"""
        today = timezone.now().date()
        report_data = []
        totals = {'total_items': 0, 'unclaimed_total': 0, 'dotace': 0, 'final_price': 0}
        grouping = 'day'
        
        if not form.is_valid():
            return report_data, totals, grouping
        
        period = form.cleaned_data['period']
        grouping = form.cleaned_data.get('grouping', 'day')
        
        # výpočet období
        if period == 'today':
            date_from = date_to = today
        elif period == 'yesterday':
            date_from = date_to = today - timedelta(days=1)
        elif period == 'week':
            date_from = date_to = today - timedelta(days=7)
        elif period == 'month':
            date_from = date_to = today.replace(day=1) - timedelta(days=1)
        elif period == 'current_month':
            date_from = today.replace(day=1)
            date_to = today
        elif period == 'year':
            date_from = today.replace(month=1, day=1)
            date_to = today
        else:  # custom
            date_from = form.cleaned_data.get('date_from')
            date_to = form.cleaned_data.get('date_to')
        
        # Queryset pro OrderItems
        queryset = OrderItem.objects.filter(
            order__status__in=['vydano', 'nevyzvednuto'],
            vydano=True
        ).select_related('order__user', 'menu_item__jidlo').order_by('order__datum_vydeje', 'order__user__first_name')
        
        if date_from and date_to:
            queryset = queryset.filter(order__datum_vydeje__range=[date_from, date_to])
        
        if form.cleaned_data['group']:
            queryset = queryset.filter(order__user__groups=form.cleaned_data['group'])
        
        if form.cleaned_data['customer']:
            queryset = queryset.filter(order__user=form.cleaned_data['customer'])
        
        for item in queryset:
            jidlo_nazev = getattr(item.menu_item.jidlo, 'nazev', 'N/A')
            jidlo_cena = getattr(item.menu_item.jidlo, 'cena', 0)
            
            unclaimed_total = item.quantity * jidlo_cena
            dotace = (jidlo_cena - item.cena) * item.quantity
            final_price = item.quantity * item.cena
            
            row = {
                'user': item.order.user.get_full_name(),
                'osobni_cislo': item.order.user.osobni_cislo or '',
                'identifikacni_medium': item.order.user.identifikacni_medium or '',
                'date': item.order.datum_vydeje,
                'status': item.order.status,
                'jidlo_nazev': jidlo_nazev,
                'quantity': item.quantity,
                'unclaimed_total': round(unclaimed_total, 2),
                'dotace': round(dotace, 2),
                'final_price': round(final_price, 2),
            }
            report_data.append(row)
            
            totals['total_items'] += item.quantity
            totals['unclaimed_total'] += unclaimed_total
            totals['dotace'] += dotace
            totals['final_price'] += final_price
        
        if grouping == 'total':
            grouped = defaultdict(
                lambda: {
                    'unclaimed_total': 0, 
                    'dotace': 0, 
                    'final_price': 0, 
                    'items_count': 0,
                    'items_list': [],
                    'osobni_cislo': '', 
                    'identifikacni_medium': ''
                }
            )
            
            for row in report_data:
                key = row['user']
                grouped[key]['unclaimed_total'] += row['unclaimed_total']
                grouped[key]['dotace'] += row['dotace']
                grouped[key]['final_price'] += row['final_price']
                grouped[key]['items_count'] += row['quantity']
                grouped[key]['items_list'].append(f"{row['jidlo_nazev']} ({row['quantity']}x)")
                
                if not grouped[key]['osobni_cislo']:
                    grouped[key]['osobni_cislo'] = row['osobni_cislo']
                if not grouped[key]['identifikacni_medium']:
                    grouped[key]['identifikacni_medium'] = row['identifikacni_medium']
            
            report_data = [{
                'user': user,
                'osobni_cislo': data['osobni_cislo'],
                'identifikacni_medium': data['identifikacni_medium'],
                'items_count': data['items_count'],
                'items_names': ', '.join(data['items_list']),
                'unclaimed_total': round(data['unclaimed_total'], 2),
                'dotace': round(data['dotace'], 2),
                'final_price': round(data['final_price'], 2),
                'grouped': True
            } for user, data in sorted(grouped.items(), key=lambda x: x[1]['final_price'], reverse=True)]

        # NOVÉ: Fulltext search
        search_query = form.cleaned_data.get('search', '').strip()
        if search_query:
            filtered_data = []
            for row in report_data:
                searchable_text = ' '.join([
                    str(row.get('user', '')),
                    str(row.get('osobni_cislo', '')),
                    str(row.get('identifikacni_medium', '')),
                    str(row.get('date', '')),
                    str(row.get('status', '')),
                    str(row.get('jidlo_nazev', '')),
                    str(row.get('quantity', '')),
                    str(row.get('items_names', '')),
                    str(row.get('unclaimed_total', '')),
                    str(row.get('dotace', '')),
                    str(row.get('final_price', '')),
                ]).lower()
                
                if search_query.lower() in searchable_text:
                    filtered_data.append(row)
            
            report_data = filtered_data
        
        return report_data, totals, grouping


    def get_food_types_report(self, form):
        """NOVÝ: Report pro druhy jídel - detailní výpis s jmény zákazníků a názvy jídel"""
        today = timezone.now().date()
        report_data = []
        totals = {'total_portions': 0, 'unclaimed_total': 0, 'dotace': 0, 'final_price': 0}
        grouping = 'day'
        
        if not form.is_valid():
            return report_data, totals, grouping
        
        period = form.cleaned_data['period']
        grouping = form.cleaned_data.get('grouping', 'day')
        
        # výpočet období
        if period == 'today':
            date_from = date_to = today
        elif period == 'yesterday':
            date_from = date_to = today - timedelta(days=1)
        elif period == 'week':
            date_from = date_to = today - timedelta(days=7)
        elif period == 'month':
            date_from = date_to = today.replace(day=1) - timedelta(days=1)
        elif period == 'current_month':
            date_from = today.replace(day=1)
            date_to = today
        elif period == 'year':
            date_from = today.replace(month=1, day=1)
            date_to = today
        else:  # custom
            date_from = form.cleaned_data.get('date_from')
            date_to = form.cleaned_data.get('date_to')
        
        # Základní queryset - detailní položky
        queryset = OrderItem.objects.filter(
            order__status__in=['vydano', 'nevyzvednuto'],
            vydano=True
        ).select_related('order__user', 'menu_item__jidlo', 'menu_item__druh_jidla').order_by(
            'order__datum_vydeje', 'order__user__first_name', 'menu_item__druh_jidla__nazev'
        )
        
        if date_from and date_to:
            queryset = queryset.filter(order__datum_vydeje__range=[date_from, date_to])
        
        if form.cleaned_data['group']:
            queryset = queryset.filter(order__user__groups=form.cleaned_data['group'])
        
        if form.cleaned_data['customer']:
            queryset = queryset.filter(order__user=form.cleaned_data['customer'])
        
        # Filtr podle druhů jídel
        if form.cleaned_data.get('food_types'):
            queryset = queryset.filter(menu_item__druh_jidla__in=form.cleaned_data['food_types'])
        
        # Procházej jednotlivé položky
        if grouping == 'day':
            for item in queryset:
                jidlo_nazev = getattr(item.menu_item.jidlo, 'nazev', 'N/A')
                jidlo_cena = getattr(item.menu_item.jidlo, 'cena', 0)
                druh_jidla = getattr(item.menu_item.druh_jidla, 'nazev', 'Neurčeno')
                
                unclaimed_total = item.quantity * jidlo_cena
                dotace = (jidlo_cena - item.cena) * item.quantity
                final_price = item.quantity * item.cena
                
                row = {
                    'date': item.order.datum_vydeje,
                    'user': item.order.user.get_full_name(),
                    'food_type': druh_jidla,
                    'food_name': jidlo_nazev,
                    'quantity': item.quantity,
                    'unclaimed_total': round(unclaimed_total, 2),
                    'dotace': round(dotace, 2),
                    'final_price': round(final_price, 2),
                }
                report_data.append(row)
                
                totals['total_portions'] += item.quantity
                totals['unclaimed_total'] += unclaimed_total
                totals['dotace'] += dotace
                totals['final_price'] += final_price
        
        else:  # grouping == 'total' - seskupení podle zákazníka + druhu + jídla
            grouped = defaultdict(
                lambda: {
                    'quantity': 0,
                    'unclaimed_total': 0,
                    'dotace': 0,
                    'final_price': 0,
                }
            )
            
            for item in queryset:
                jidlo_nazev = getattr(item.menu_item.jidlo, 'nazev', 'N/A')
                jidlo_cena = getattr(item.menu_item.jidlo, 'cena', 0)
                druh_jidla = getattr(item.menu_item.druh_jidla, 'nazev', 'Neurčeno')
                user_name = item.order.user.get_full_name()
                
                # Klíč: zákazník + druh jídla + název jídla
                key = (user_name, druh_jidla, jidlo_nazev)
                
                unclaimed_total = item.quantity * jidlo_cena
                dotace = (jidlo_cena - item.cena) * item.quantity
                final_price = item.quantity * item.cena
                
                grouped[key]['quantity'] += item.quantity
                grouped[key]['unclaimed_total'] += unclaimed_total
                grouped[key]['dotace'] += dotace
                grouped[key]['final_price'] += final_price
            
            # Převeď na list
            for (user_name, druh_jidla, jidlo_nazev), data in sorted(
                grouped.items(), 
                key=lambda x: x[1]['final_price'], 
                reverse=True
            ):
                row = {
                    'user': user_name,
                    'food_type': druh_jidla,
                    'food_name': jidlo_nazev,
                    'quantity': data['quantity'],
                    'unclaimed_total': round(data['unclaimed_total'], 2),
                    'dotace': round(data['dotace'], 2),
                    'final_price': round(data['final_price'], 2),
                    'grouped': True
                }
                report_data.append(row)
                
                totals['total_portions'] += data['quantity']
                totals['unclaimed_total'] += data['unclaimed_total']
                totals['dotace'] += data['dotace']
                totals['final_price'] += data['final_price']

        # Fulltext search
        search_query = form.cleaned_data.get('search', '').strip()
        if search_query:
            filtered_data = []
            for row in report_data:
                searchable_text = ' '.join([
                    str(row.get('date', '')),
                    str(row.get('user', '')),
                    str(row.get('food_type', '')),
                    str(row.get('food_name', '')),
                    str(row.get('quantity', '')),
                    str(row.get('unclaimed_total', '')),
                    str(row.get('dotace', '')),
                    str(row.get('final_price', '')),
                ]).lower()
                
                if search_query.lower() in searchable_text:
                    filtered_data.append(row)
            
            report_data = filtered_data
        
        return report_data, totals, grouping


    def dashboard_view(self, request):
        active_report = request.GET.get('report', 'castky')
        export_type = request.GET.get('export')
        form = ReportForm(request.GET or None)
        
        # Rozhodnutí, který report použít
        if active_report == 'polozky':
            report_data, totals, grouping = self.get_order_items_report(form)
            report_type = 'items'
        elif active_report == 'druhy-jidel':  # NOVÉ
            report_data, totals, grouping = self.get_food_types_report(form)
            report_type = 'food_types'
        else:
            report_data, totals, grouping = self.get_report_calculations(form)
            report_type = 'amounts'

        if export_type and report_data:
            filename_base = f"report_{active_report}_{timezone.now().strftime('%Y%m%d')}"

            # CSV Export
            if export_type == 'csv':
                response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
                response['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
                writer = csv.writer(response, delimiter=';')
                
                if report_type == 'food_types':
                    headers = []
                    if grouping == 'day': headers += ['Datum']
                    headers += ['Zákazník', 'Druh jídla', 'Název jídla', 'Počet ks', 'Plná částka', 'Dotace', 'K platbě']
                else:
                    headers = ['Zákazník', 'Osobní číslo', 'ID médium']
                    if grouping == 'day': headers += ['Datum', 'Stav']
                    if report_type == 'items':
                        headers += (['Jídlo', 'Počet ks'] if grouping == 'day' else ['Počet položek', 'Jídla'])
                    headers += ['Plná částka', 'Dotace', 'K platbě']
                    if report_type == 'amounts' and grouping == 'total': headers += ['Počet objednávek']
                
                writer.writerow(headers)
                
                for row in report_data:
                    if report_type == 'food_types':
                        line = []
                        if grouping == 'day': line += [row['date'].strftime('%d.%m.%Y')]
                        line += [row['user'], row['food_type'], row['food_name'], row['quantity'],
                                str(row['unclaimed_total']).replace('.', ','),
                                str(row['dotace']).replace('.', ','),
                                str(row['final_price']).replace('.', ',')]
                    else:
                        line = [row['user'], row['osobni_cislo'], row['identifikacni_medium']]
                        if grouping == 'day': line += [row['date'].strftime('%d.%m.%Y'), row['status']]
                        if report_type == 'items':
                            line += ([row.get('jidlo_nazev', ''), row.get('quantity', 0)] if grouping == 'day' 
                                    else [row.get('items_count', 0), row.get('items_names', '')])
                        line += [str(row['unclaimed_total']).replace('.', ','),
                                str(row['dotace']).replace('.', ','),
                                str(row['final_price']).replace('.', ',')]
                        if report_type == 'amounts' and grouping == 'total': line += [row.get('count', 0)]
                    writer.writerow(line)
                return response

            # XLSX Export
            elif export_type == 'xls' and OPENPYXL_AVAILABLE:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Report"
                
                if report_type == 'food_types':
                    headers = []
                    if grouping == 'day': headers += ['Datum']
                    headers += ['Zákazník', 'Druh jídla', 'Název jídla', 'Počet ks', 'Plná částka', 'Dotace', 'K platbě']
                else:
                    headers = ['Zákazník', 'Osobní číslo', 'ID médium']
                    if grouping == 'day': headers += ['Datum', 'Stav']
                    if report_type == 'items':
                        headers += (['Jídlo', 'Počet ks'] if grouping == 'day' else ['Počet položek', 'Jídla'])
                    headers += ['Plná částka', 'Dotace', 'K platbě']
                    if report_type == 'amounts' and grouping == 'total': headers += ['Počet objednávek']
                
                ws.append(headers)
                
                for row in report_data:
                    if report_type == 'food_types':
                        line = []
                        if grouping == 'day': line += [row['date'].strftime('%d.%m.%Y')]
                        line += [row['user'], row['food_type'], row['food_name'], row['quantity'],
                                row['unclaimed_total'], row['dotace'], row['final_price']]
                    else:
                        line = [row['user'], row['osobni_cislo'], row['identifikacni_medium']]
                        if grouping == 'day': line += [row['date'].strftime('%d.%m.%Y'), row['status']]
                        if report_type == 'items':
                            line += ([row.get('jidlo_nazev', ''), row.get('quantity', 0)] if grouping == 'day' 
                                    else [row.get('items_count', 0), row.get('items_names', '')])
                        line += [row['unclaimed_total'], row['dotace'], row['final_price']]
                        if report_type == 'amounts' and grouping == 'total': line += [row.get('count', 0)]
                    ws.append(line)
                
                # Footer XLSX
                footer = ['CELKEM']
                col_offset = len(headers) - 3
                footer += [''] * (col_offset - 1)
                footer += [totals.get('total_portions', totals.get('total_items', '')), 
                          totals['unclaimed_total'], totals['dotace'], totals['final_price']]
                ws.append(footer)
                ws.cell(row=ws.max_row, column=1).font = Font(bold=True)
                
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
                wb.save(response)
                return response

            # PDF Export
            # PDF Export
            elif export_type == 'pdf':
                response = HttpResponse(content_type='application/pdf')
                response['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
                font_dir = os.path.join(settings.BASE_DIR, 'static', 'fonts')
                
                try:
                    pdfmetrics.registerFont(TTFont('DejaVu', os.path.join(font_dir, 'DejaVuSans.ttf')))
                    pdfmetrics.registerFont(TTFont('DejaVu-Bold', os.path.join(font_dir, 'DejaVuSans-Bold.ttf')))
                    font_name, font_bold = 'DejaVu', 'DejaVu-Bold'
                except:
                    font_name, font_bold = "Helvetica", "Helvetica-Bold"
                
                # Barvy z projektu
                color_green = colors.HexColor('#28a745') # Tmavě zelená
                color_orange = colors.HexColor('#fd7e14') # Oranžová
                
                # Správný čas pro Prahu
                current_time = timezone.localtime(timezone.now())
                
                doc = SimpleDocTemplate(
                    response,
                    pagesize=landscape(A4),
                    rightMargin=0.5*cm,
                    leftMargin=0.5*cm,
                    topMargin=1.5*cm,
                    bottomMargin=1.5*cm
                )
                
                styles = getSampleStyleSheet()
                style_heading = ParagraphStyle(
                    'CustomHeading', parent=styles['Heading1'],
                    fontName=font_bold, fontSize=16, textColor=color_orange
                )
                style_normal = ParagraphStyle(
                    'CustomNormal', parent=styles['Normal'],
                    fontName=font_name, fontSize=8, leading=10
                )
                
                story = []
                
                # Nadpis
                title_text = {
                    'items': 'Přehled prodaných jídel',
                    'amounts': 'Přehled objednávek',
                    'food_types': 'Přehled podle druhu jídel'
                }.get(report_type, 'Report')
                
                story.append(Paragraph(title_text, style_heading))
                story.append(Spacer(1, 0.3*cm))
                
                # --- FILTRY DO PDF ---
                period = request.GET.get('period', 'vše')
                grouping_display = 'Denní' if request.GET.get('grouping') == 'day' else 'Celkem'
                period_display = {
                    'today': 'Dnes', 'yesterday': 'Včera', 'week': 'Týden', 
                    'month': 'Minulý měsíc', 'current_month': 'Aktuální měsíc', 
                    'year': 'Rok', 'custom': 'Vlastní'
                }.get(period, period)
                
                # Zákazník
                customer_id = request.GET.get('customer')
                customer_name = User.objects.filter(id=customer_id).first().get_full_name() if customer_id else 'Všichni'
                
                # Druhy jídel
                food_types_ids = request.GET.getlist('food_types')
                food_types_str = ", ".join([obj.nazev for obj in DruhJidla.objects.filter(id__in=food_types_ids)]) if food_types_ids else "Všechny"
                
                # Vyhledávání
                search_query = request.GET.get('search', '').strip()
                search_display = f'"{search_query}"' if search_query else "Žádné"

                filter_text = f"""
                <b>Období:</b> {period_display} | <b>Seskupení:</b> {grouping_display} | 
                <b>Zákazník:</b> {customer_name} | <b>Druhy jídel:</b> {food_types_str} | 
                <b>Hledat:</b> {search_display}
                """
                story.append(Paragraph(filter_text, style_normal))
                story.append(Spacer(1, 0.5*cm))
                
                # Definice šířek (celkem cca 28 cm pro landscape A4)
                if report_type == 'food_types':
                    if grouping == 'day':
                        headers = ['Datum', 'Zákazník', 'Druh jídla', 'Název jídla', 'Ks', 'Plná', 'Dotace', 'K platbě']
                        col_widths = [2.5*cm, 4.5*cm, 3.5*cm, 8.0*cm, 1.2*cm, 2.8*cm, 2.8*cm, 2.8*cm]
                    else:
                        headers = ['Zákazník', 'Druh jídla', 'Název jídla', 'Ks', 'Plná', 'Dotace', 'K platbě']
                        col_widths = [6.0*cm, 4.0*cm, 9.0*cm, 1.5*cm, 2.5*cm, 2.5*cm, 2.5*cm]
                else:
                    if grouping == 'day':
                        headers = ['Zákazník', 'Osobní č.', 'ID médium', 'Datum', 'Stav', 'Plná', 'Dotace', 'K platbě']
                        col_widths = [5.0*cm, 2.5*cm, 3.0*cm, 3.0*cm, 2.5*cm, 3.8*cm, 3.8*cm, 3.8*cm]
                    else:
                        headers = ['Zákazník', 'Osobní č.', 'ID médium', 'Počet ks', 'Plná', 'Dotace', 'K platbě']
                        col_widths = [7.5*cm, 4.0*cm, 4.0*cm, 2.5*cm, 3.3*cm, 3.3*cm, 3.3*cm]
                
                table_data = [headers]
                for row in report_data:
                    if report_type == 'food_types':
                        line = [row['date'].strftime('%d.%m.%Y')] if grouping == 'day' else []
                        line += [row['user'], row['food_type'], row['food_name'], str(row['quantity']), 
                                f"{row['unclaimed_total']:.2f} Kč", f"{row['dotace']:.2f} Kč", f"{row['final_price']:.2f} Kč"]
                        table_data.append(line)
                    else:
                        count_val = f"{row.get('count', row.get('items_count', 0))} ks"
                        if grouping == 'day':
                            table_data.append([row['user'], row['osobni_cislo'], row['identifikacni_medium'], row['date'].strftime('%d.%m.%Y'), row['status'], f"{row['unclaimed_total']:.2f} Kč", f"{row['dotace']:.2f} Kč", f"{row['final_price']:.2f} Kč"])
                        else:
                            table_data.append([row['user'], row['osobni_cislo'], row['identifikacni_medium'], count_val, f"{row['unclaimed_total']:.2f} Kč", f"{row['dotace']:.2f} Kč", f"{row['final_price']:.2f} Kč"])

                # Footer
                if report_type == 'food_types':
                    footer = ['CELKEM'] + ([''] if grouping == 'day' else []) + ['', '', str(totals['total_portions']), f"{totals['unclaimed_total']:.2f} Kč", f"{totals['dotace']:.2f} Kč", f"{totals['final_price']:.2f} Kč"]
                else:
                    footer = ['CELKEM'] + (['', '', '', ''] if grouping == 'day' else ['', '', '']) + [f"{totals['unclaimed_total']:.2f} Kč", f"{totals['dotace']:.2f} Kč", f"{totals['final_price']:.2f} Kč"]
                table_data.append(footer)

                table = Table(table_data, colWidths=col_widths)
                table.setStyle(TableStyle([
                    ('FONTNAME', (0, 0), (-1, -1), font_name),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('BACKGROUND', (0, 0), (-1, 0), color_green),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), font_bold),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#fff3e0')),
                    ('FONTNAME', (0, -1), (-1, -1), font_bold),
                    ('LINEABOVE', (0, -1), (-1, -1), 1, color_orange),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('ALIGN', (-3, 1), (-1, -1), 'RIGHT'), # Částky vpravo
                ]))
                
                story.append(table)
                story.append(Spacer(1, 0.6*cm))
                story.append(Paragraph(f"<i>Vygenerováno: {current_time.strftime('%d.%m.%Y %H:%M:%S')} (Praha)</i>", style_normal))
                
                doc.build(story)
                return response



        context = {
            **self.admin_site.each_context(request),
            'form': form,
            'report_data': report_data,
            'totals': totals,
            'active_report': active_report,
            'report_type': report_type,
            'grouping': grouping,
            'reports': [
                {'id': 'castky', 'title': 'Částky objednávek', 'icon': '💰'},
                
                {'id': 'druhy-jidel', 'title': 'Druhy jídel', 'icon': '📊'},  # NOVÉ
            ],
            'opts': self.model._meta,
            'title': 'Reporty Dashboard',
        }

        return render(request, 'admin/reporty/dashboard.html', context)